"""
Tender Summary Generator — extracts text from downloaded PDFs/XLS
and produces a structured one-pager summary.

Strategy:
1. Extract ALL text from PDF (page by page)
2. Use TOC/headings to identify sections (Scope of Work, Eligibility, etc.)
3. Extract structured fields from portal metadata + PDF content
4. If tender value missing, estimate as 20×EMD
"""

import json
import logging
import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


def extract_pdf_text(pdf_path: str, max_pages: int = 60) -> Tuple[str, List[str]]:
    """
    Extract text from PDF. Returns (full_text, per_page_list).
    """
    pages = []
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(pdf_path)
        for i, page in enumerate(reader.pages[:max_pages]):
            try:
                t = page.extract_text() or ""
                pages.append(t)
            except Exception as e:
                logger.warning(f"Page {i} extraction failed: {e}")
                pages.append("")
    except Exception as e:
        logger.error(f"PDF read failed for {pdf_path}: {e}")
    return "\n\n".join(pages), pages


def extract_xls_text(xls_path: str) -> str:
    """Extract text from XLS/XLSX (best effort)."""
    try:
        if xls_path.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(xls_path, read_only=True)
                lines = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                return "\n".join(lines)
            except ImportError:
                pass
        try:
            import xlrd
            wb = xlrd.open_workbook(xls_path)
            lines = []
            for sheet in wb.sheets():
                for row_idx in range(sheet.nrows):
                    lines.append(" | ".join(str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)))
            return "\n".join(lines)
        except ImportError:
            pass
        return f"[XLS file: {Path(xls_path).name}]"
    except Exception as e:
        logger.error(f"XLS extraction failed: {e}")
        return ""


def _find_section(pages: List[str], headings: List[str], max_chars: int = 3000) -> str:
    """
    Find a section in PDF pages by looking for heading keywords.
    Returns the text from that heading until the next major heading.
    """
    full = "\n\n".join(pages)
    
    for heading in headings:
        # Try to find the heading in the text
        pattern = re.compile(
            r'(?:^|\n)\s*(?:\d+\.?\s*)?' + re.escape(heading) + r'\s*\n(.*?)(?=\n\s*(?:\d+\.?\s*)?(?:TECHNICAL BID|COMMERCIAL BID|GENERAL GUIDE|SUBMISSION OF BID|BID OPENING|AWARD|ANNEXURE|SCHEDULE|CONTENTS)\s*\n|\Z)',
            re.IGNORECASE | re.DOTALL
        )
        m = pattern.search(full)
        if m:
            text = m.group(1).strip()
            if len(text) > 50:  # meaningful content
                return text[:max_chars]
    
    # Fallback: search page by page for heading keyword
    for heading in headings:
        for i, page in enumerate(pages):
            if heading.lower() in page.lower():
                # Grab this page + next page
                section = page
                if i + 1 < len(pages):
                    section += "\n" + pages[i + 1]
                # Try to extract from the heading onwards
                idx = section.lower().find(heading.lower())
                if idx >= 0:
                    text = section[idx + len(heading):].strip().lstrip(':').strip()
                    if len(text) > 50:
                        return text[:max_chars]
    
    return ""


def _find_eligibility(pages: List[str], full_text: str) -> str:
    """Extract eligibility criteria from PDF."""
    # Look for specific eligibility sections
    eligibility_markers = [
        "Documents required to be submitted for Qualification",
        "Eligibility Criteria",
        "Qualification Criteria", 
        "Pre-Qualification",
        "Eligible Bidders",
        "Eligibility",
    ]
    
    for marker in eligibility_markers:
        for i, page in enumerate(pages):
            if marker.lower() in page.lower():
                idx = page.lower().find(marker.lower())
                text = page[idx:].strip()
                # Include next page too for continuity
                if i + 1 < len(pages):
                    text += "\n" + pages[i + 1]
                # Clean up and truncate
                text = _clean_text(text, max_len=1500)
                if len(text) > 80:
                    return text
    
    # Fallback: look for registration/class requirements
    reg_patterns = [
        r'(?:Registration\s+in\s+.+?(?:Class|Category).+?)(?:\n\n|\n\d+\.)',
        r'(?:The\s+(?:bidder|tenderer)\s+(?:should|shall|must).+?(?:registered|experience|qualification).+?)(?:\n\n)',
    ]
    for pat in reg_patterns:
        m = re.search(pat, full_text, re.IGNORECASE | re.DOTALL)
        if m:
            return _clean_text(m.group(0), max_len=1500)
    
    return ""


def _find_scope(pages: List[str], full_text: str) -> str:
    """Extract scope of work from PDF."""
    scope_markers = [
        "Scope of Work",
        "Scope Of Work", 
        "SCOPE OF WORK",
        "Description of Work",
        "Brief Scope",
    ]
    
    # Strategy: find the page where scope CONTENT starts (not TOC reference).
    # Skip pages that look like TOC (contain multiple "Page No" or numbered section listings).
    collected = []
    found_start = False
    
    for i, page in enumerate(pages):
        page_lower = page.lower()
        
        if not found_start:
            for marker in scope_markers:
                if marker.lower() in page_lower:
                    # Skip TOC pages — they have many section references like "13-18", "Page No"
                    if re.search(r'\d+-\d+.*\n.*\d+-\d+', page) and "page no" in page_lower:
                        continue
                    idx = page_lower.find(marker.lower())
                    remaining = page[idx + len(marker):].strip()
                    # Must have substantial content (not just a heading)
                    if len(remaining) < 100:
                        continue
                    # Prefer pages where scope is a standalone heading (near top of page)
                    # Skip if "scope of work" appears mid-page as part of another section
                    lines_before = page[:idx].strip().split('\n')
                    # If there are many content lines before the marker, it's likely embedded
                    # in another section (e.g., Special Terms). Prefer pages where it's near top.
                    non_empty_before = [l for l in lines_before if l.strip() and not re.match(r'^\s*\d{1,3}\s*$', l)]
                    if len(non_empty_before) > 3:
                        # This page has lots of content before "Scope of Work" — likely not the right page
                        # But keep as fallback
                        if not collected:
                            # Store as fallback, keep looking
                            pass
                        continue
                    collected.append(remaining)
                    found_start = True
                    break
        elif found_start:
            # Check if we've hit a new major section
            if any(h.lower() in page_lower for h in [
                "technical bid", "commercial bid", "general guide",
                "submission of bid", "bid opening", "schedule g"
            ]):
                break
            collected.append(page.strip())
            # Limit to ~4 pages of scope
            if len(collected) >= 5:
                break
    
    if collected:
        return _clean_text("\n\n".join(collected), max_len=2000)
    
    return ""


def _clean_text(text: str, max_len: int = 1500) -> str:
    """Clean extracted text: remove excess whitespace, page numbers, etc."""
    # Remove standalone page numbers
    text = re.sub(r'^\s*\d{1,3}\s*$', '', text, flags=re.MULTILINE)
    # Collapse multiple newlines
    text = re.sub(r'\n{3,}', '\n\n', text)
    # Collapse multiple spaces
    text = re.sub(r'  +', ' ', text)
    text = text.strip()
    if len(text) > max_len:
        # Try to cut at sentence boundary
        cut = text[:max_len].rfind('.')
        if cut > max_len * 0.7:
            text = text[:cut + 1]
        else:
            text = text[:max_len] + "..."
    return text


def _parse_amount(text: str) -> Optional[float]:
    """Parse amount string to float."""
    if not text:
        return None
    cleaned = re.sub(r'[₹Rs.\s]', '', text).replace(',', '')
    try:
        return float(cleaned)
    except ValueError:
        return None


def _format_amount(amount: float) -> str:
    """Format in Indian notation."""
    if amount >= 1_00_00_000:
        return f"₹ {amount / 1_00_00_000:.2f} Cr"
    elif amount >= 1_00_000:
        return f"₹ {amount / 1_00_000:.2f} Lakh"
    else:
        return f"₹ {amount:,.0f}"


def _find_in_details(details: Dict, keys: list) -> Optional[str]:
    """Find value in portal-scraped details."""
    for k in keys:
        for dk, dv in details.items():
            if k.lower() in dk.lower() and dv and len(dv) < 500 and dv.strip() not in ["", "NA", "N/A", "--"]:
                return dv.strip()
    return None


def _find_in_text(text: str, patterns: list) -> Optional[str]:
    """Find first matching pattern."""
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            result = m.group(1) if m.lastindex else m.group(0)
            return result.strip()
    return None


def generate_detailed_summary(tender_dir: str, portal_details: Dict = None) -> Dict:
    """
    Generate a one-pager summary from downloaded tender documents.
    
    Extracts:
    - Tender Details, Publishing Agency, Published Date, Last Date,
      Pre-Bid Date, EMD, Estimated Value (or 20×EMD*), 
      Brief Scope of Work, Eligibility Criteria, JV Allowed
    """
    tender_path = Path(tender_dir)
    details = portal_details or {}
    
    # Collect text from all documents
    all_pages = []  # per-page text from PDFs
    full_text = ""
    boq_text = ""
    doc_files = []
    
    for search_dir in [tender_path / "extracted", tender_path / "downloads", tender_path]:
        if not search_dir.exists():
            continue
        for f in sorted(search_dir.iterdir()):
            if f.is_file():
                if f.suffix.lower() == '.pdf':
                    doc_files.append(f.name)
                    text, pages = extract_pdf_text(str(f))
                    all_pages.extend(pages)
                    full_text += f"\n\n{text}"
                elif f.suffix.lower() in ['.xls', '.xlsx']:
                    doc_files.append(f.name)
                    boq_text += extract_xls_text(str(f))
    
    if not full_text.strip() and not details:
        return {"error": "No document text could be extracted"}
    
    summary = {}
    
    # --- Tender Details ---
    summary["tender_title"] = (
        _find_in_details(details, ["Title", "Work Description", "Tender Name"]) or
        _find_in_text(full_text, [
            r"(?:Tender document for|Name of Work|Work Name)[:\s]+(.+?)(?:\n)",
        ]) or
        "See attached documents"
    )
    
    summary["tender_id"] = _find_in_details(details, ["Tender ID", "Tender Reference Number", "NIT No"])
    
    summary["tender_type"] = (
        _find_in_details(details, ["Tender Type", "Bid Type"]) or
        _find_in_text(full_text, [r"((?:Two|Single)\s*(?:Bid|Cover|Packet)\s*(?:System)?)"])
    )
    
    # --- Publishing Agency ---
    summary["publishing_agency"] = (
        _find_in_details(details, ["Organisation Chain", "Organization", "Department"]) or
        _find_in_text(full_text, [
            r"(?:OFFICE OF THE\s+)(.+?)(?:\n\n)",
        ])
    )
    
    # --- Dates ---
    summary["published_date"] = _find_in_details(details, ["Published Date", "Publication Date"])
    
    summary["last_date"] = (
        _find_in_details(details, ["Bid Submission End Date", "Last Date", "Closing Date"]) or
        _find_in_text(full_text, [
            r"(?:Last Date|Closing Date|Bid Submission End)[:\s]+(.+?)(?:\n)",
        ])
    )
    
    summary["pre_bid_date"] = (
        _find_in_details(details, ["Pre Bid Meeting Date", "Pre-Bid Meeting"]) or
        _find_in_text(full_text, [
            r"(?:Pre[\s-]*Bid[\s]*Meeting.*?date)[:\s]+(.+?)(?:\n)",
        ])
    )
    
    summary["bid_opening_date"] = (
        _find_in_details(details, ["Bid Opening Date", "Technical Bid Opening"]) or
        _find_in_text(full_text, [
            r"(?:Technical Bid Opening|Bid Opening Date)[:\s]+(.+?)(?:\n)",
        ])
    )
    
    # --- EMD ---
    emd_str = (
        _find_in_details(details, ["EMD Amount", "Earnest Money", "EMD"]) or
        _find_in_text(full_text, [
            r"(?:EMD|Earnest Money|Earnest money)[^0-9]*?(?:Rs\.?\s*)?(\d[\d,]+(?:\.\d+)?)",
        ])
    )
    summary["emd"] = emd_str
    
    # --- Estimated Value ---
    tender_value_str = (
        _find_in_details(details, ["Tender Value", "Estimated Cost", "Estimated Value"]) or
        _find_in_text(full_text, [
            r"(?:Estimated\s*(?:Cost|Value)|Tender Value)[:\s]*(?:Rs\.?|₹)?\s*([\d,]+(?:\.\d+)?)",
        ])
    )
    
    if tender_value_str and tender_value_str.strip() not in ["", "0", "Refer Document", "NA"]:
        summary["estimated_value"] = tender_value_str
        summary["estimated_value_note"] = None
    else:
        emd_amount = _parse_amount(emd_str) if emd_str else None
        if emd_amount and emd_amount > 0:
            estimated = emd_amount * 20
            summary["estimated_value"] = _format_amount(estimated) + " *"
            summary["estimated_value_note"] = "* Estimated as 20× EMD (actual value not available in documents)"
        else:
            summary["estimated_value"] = "Not Available"
            summary["estimated_value_note"] = None
    
    # --- Scope of Work (page-aware extraction) ---
    scope = _find_scope(all_pages, full_text)
    if not scope:
        scope = (
            _find_in_details(details, ["Work Description", "Title"]) or
            summary.get("tender_title", "Refer tender documents")
        )
    summary["scope_of_work"] = scope
    
    # --- Eligibility Criteria (page-aware extraction) ---
    eligibility = _find_eligibility(all_pages, full_text)
    if not eligibility:
        eligibility = (
            _find_in_details(details, ["NDA/Pre Qualification", "Eligibility"]) or
            "Refer tender documents for detailed eligibility criteria"
        )
    summary["eligibility_criteria"] = eligibility
    
    # --- JV Allowed ---
    jv_text = full_text.lower()
    if re.search(r"joint\s*venture.*(?:not\s*(?:allowed|permitted|acceptable))", jv_text):
        summary["jv_allowed"] = "❌ No — Joint Venture not allowed"
    elif re.search(r"joint\s*venture.*(?:allowed|permitted|acceptable|may\s*(?:bid|participate))", jv_text):
        summary["jv_allowed"] = "✅ Yes — Joint Venture allowed"
    elif re.search(r"(?:consortium|jv).*(?:allowed|permitted)", jv_text):
        summary["jv_allowed"] = "✅ Yes — JV/Consortium allowed"
    elif re.search(r"(?:consortium|jv|joint\s*venture).*(?:not\s*allowed|prohibited)", jv_text):
        summary["jv_allowed"] = "❌ No — JV/Consortium not allowed"
    elif "joint venture" in jv_text or " jv " in jv_text:
        summary["jv_allowed"] = "⚠️ JV mentioned in documents — check details"
    else:
        summary["jv_allowed"] = "ℹ️ Not specified in available documents"
    
    # --- Tender Fee ---
    summary["tender_fee"] = (
        _find_in_details(details, ["Tender Fee", "Document Fee", "Cost of Tender"]) or
        _find_in_text(full_text, [
            r"(?:Cost of tender document|Tender Fee|Document Fee)[:\s]*(?:Rs\.?|₹)?\s*([\d,]+(?:\.\d+)?(?:\s*\+\s*\d+%?\s*GST)?)",
        ])
    )
    
    # --- Documents ---
    summary["documents"] = doc_files
    
    return summary
