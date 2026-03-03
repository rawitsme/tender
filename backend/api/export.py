"""Export endpoints — Excel/CSV download of search results."""

import io
from datetime import datetime
from typing import Optional, List
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.services.search import search_tenders
from backend.schemas.tender import TenderSearchRequest

router = APIRouter(prefix="/export", tags=["export"])


@router.post("/xlsx")
async def export_xlsx(req: TenderSearchRequest, db: AsyncSession = Depends(get_db)):
    """Export search results as Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Fetch up to 1000 results
    req.page_size = min(req.page_size, 1000)
    tenders, total = await search_tenders(
        db, query=req.query, states=req.states, sources=req.sources,
        categories=req.categories, departments=req.departments,
        status=req.status, min_value=req.min_value, max_value=req.max_value,
        closing_within=req.closing_within, department_search=req.department,
        category_search=req.category, page=1, page_size=1000,
        sort_by=req.sort_by, sort_order=req.sort_order,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenders"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )

    headers = [
        "Tender ID", "Title", "Source", "State", "Department", "Organization",
        "Category", "Type", "Estimated Value (₹)", "EMD (₹)", "Doc Fee (₹)",
        "Published", "Bid Opens", "Bid Closes", "Status",
        "Contact Person", "Contact Email", "Contact Phone", "Source URL"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, t in enumerate(tenders, 2):
        ws.cell(row=row_idx, column=1, value=t.tender_id)
        ws.cell(row=row_idx, column=2, value=t.title[:200])
        ws.cell(row=row_idx, column=3, value=str(t.source.value if hasattr(t.source, 'value') else t.source).upper())
        ws.cell(row=row_idx, column=4, value=t.state)
        ws.cell(row=row_idx, column=5, value=t.department)
        ws.cell(row=row_idx, column=6, value=t.organization)
        ws.cell(row=row_idx, column=7, value=t.category)
        ws.cell(row=row_idx, column=8, value=str(t.tender_type.value if hasattr(t.tender_type, 'value') else t.tender_type) if t.tender_type else None)
        ws.cell(row=row_idx, column=9, value=float(t.tender_value_estimated) if t.tender_value_estimated else None)
        ws.cell(row=row_idx, column=10, value=float(t.emd_amount) if t.emd_amount else None)
        ws.cell(row=row_idx, column=11, value=float(t.document_fee) if t.document_fee else None)
        ws.cell(row=row_idx, column=12, value=t.publication_date.strftime("%d %b %Y") if t.publication_date else None)
        ws.cell(row=row_idx, column=13, value=t.bid_open_date.strftime("%d %b %Y") if t.bid_open_date else None)
        ws.cell(row=row_idx, column=14, value=t.bid_close_date.strftime("%d %b %Y") if t.bid_close_date else None)
        ws.cell(row=row_idx, column=15, value=str(t.status.value if hasattr(t.status, 'value') else t.status).upper())
        ws.cell(row=row_idx, column=16, value=t.contact_person)
        ws.cell(row=row_idx, column=17, value=t.contact_email)
        ws.cell(row=row_idx, column=18, value=t.contact_phone)
        ws.cell(row=row_idx, column=19, value=t.source_url)

        for col in range(1, 20):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-width
    for col in range(1, 20):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.column_dimensions['B'].width = 50  # Title wider

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"tenders_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/csv")
async def export_csv(req: TenderSearchRequest, db: AsyncSession = Depends(get_db)):
    """Export search results as CSV."""
    import csv

    tenders, _ = await search_tenders(
        db, query=req.query, states=req.states, sources=req.sources,
        status=req.status, closing_within=req.closing_within,
        page=1, page_size=1000, sort_by=req.sort_by, sort_order=req.sort_order,
    )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Tender ID", "Title", "Source", "State", "Department", "Organization",
        "Estimated Value", "EMD", "Published", "Bid Closes", "Status", "Source URL"
    ])

    for t in tenders:
        writer.writerow([
            t.tender_id, t.title[:200],
            str(t.source.value if hasattr(t.source, 'value') else t.source).upper(),
            t.state, t.department, t.organization,
            float(t.tender_value_estimated) if t.tender_value_estimated else "",
            float(t.emd_amount) if t.emd_amount else "",
            t.publication_date.strftime("%d %b %Y") if t.publication_date else "",
            t.bid_close_date.strftime("%d %b %Y") if t.bid_close_date else "",
            str(t.status.value if hasattr(t.status, 'value') else t.status),
            t.source_url,
        ])

    output = buf.getvalue().encode('utf-8-sig')
    filename = f"tenders_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(output),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
