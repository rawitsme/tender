"""Download Center API — cascading filters + template-formatted XLSX export."""

import io
from datetime import datetime, date
from typing import Optional, List
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, Field

from backend.database import get_db

router = APIRouter(prefix="/download-center", tags=["Download Center"])


# ── Filter Options ──────────────────────────────────────────────────────────

@router.get("/source-categories")
async def get_source_categories(db: AsyncSession = Depends(get_db)):
    """Return grouped source categories: State Portals, Central Portals, Other."""
    result = await db.execute(text(
        "SELECT source, COUNT(*) as cnt FROM tenders WHERE status = 'ACTIVE' "
        "GROUP BY source ORDER BY cnt DESC"
    ))
    rows = result.fetchall()

    central = []
    state = []
    for src, cnt in rows:
        entry = {"id": src, "label": src.upper(), "count": cnt}
        if src.upper() in ("CPPP", "GEM"):
            central.append(entry)
        else:
            state.append(entry)

    return {
        "categories": [
            {"id": "state", "label": "State Portals", "sources": state},
            {"id": "central", "label": "Central Portals", "sources": central},
        ]
    }


@router.get("/states")
async def get_states(
    sources: Optional[str] = Query(None, description="Comma-separated source ids"),
    db: AsyncSession = Depends(get_db),
):
    """Return states with active tender counts, filtered by source."""
    where = "WHERE status = 'ACTIVE' AND state IS NOT NULL AND state != ''"
    params = {}
    if sources:
        src_list = [s.strip() for s in sources.split(",") if s.strip()]
        if src_list:
            where += " AND source = ANY(:sources)"
            params["sources"] = src_list

    result = await db.execute(
        text(f"SELECT state, COUNT(*) as cnt FROM tenders {where} GROUP BY state ORDER BY cnt DESC"),
        params,
    )
    return [{"id": row[0], "label": row[0], "count": row[1]} for row in result.fetchall()]


@router.get("/authorities")
async def get_authorities(
    sources: Optional[str] = Query(None),
    states: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Return organizations (tender authorities) with counts."""
    where = "WHERE status = 'ACTIVE' AND organization IS NOT NULL AND organization != ''"
    params = {}
    if sources:
        src_list = [s.strip() for s in sources.split(",") if s.strip()]
        if src_list:
            where += " AND source = ANY(:sources)"
            params["sources"] = src_list
    if states:
        st_list = [s.strip() for s in states.split(",") if s.strip()]
        if st_list:
            where += " AND state = ANY(:states)"
            params["states"] = st_list

    result = await db.execute(
        text(f"SELECT organization, COUNT(*) as cnt FROM tenders {where} "
             "GROUP BY organization ORDER BY cnt DESC"),
        params,
    )
    return [{"id": row[0], "label": row[0], "count": row[1]} for row in result.fetchall()]


@router.get("/departments")
async def get_departments(
    sources: Optional[str] = Query(None),
    states: Optional[str] = Query(None),
    authorities: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Return departments with counts."""
    where = "WHERE status = 'ACTIVE' AND department IS NOT NULL AND department != ''"
    params = {}
    if sources:
        src_list = [s.strip() for s in sources.split(",") if s.strip()]
        if src_list:
            where += " AND source = ANY(:sources)"
            params["sources"] = src_list
    if states:
        st_list = [s.strip() for s in states.split(",") if s.strip()]
        if st_list:
            where += " AND state = ANY(:states)"
            params["states"] = st_list
    if authorities:
        auth_list = [a.strip() for a in authorities.split(",") if a.strip()]
        if auth_list:
            where += " AND organization = ANY(:authorities)"
            params["authorities"] = auth_list

    result = await db.execute(
        text(f"SELECT department, COUNT(*) as cnt FROM tenders {where} "
             "GROUP BY department ORDER BY cnt DESC"),
        params,
    )
    return [{"id": row[0], "label": row[0], "count": row[1]} for row in result.fetchall()]


# ── Export XLSX (template format) ───────────────────────────────────────────

class DownloadRequest(BaseModel):
    sources: Optional[List[str]] = None
    states: Optional[List[str]] = None
    authorities: Optional[List[str]] = None  # organizations
    departments: Optional[List[str]] = None
    query: Optional[str] = None
    min_value: Optional[Decimal] = None
    max_value: Optional[Decimal] = None


@router.post("/preview-count")
async def preview_count(req: DownloadRequest, db: AsyncSession = Depends(get_db)):
    """Return count of tenders matching filters."""
    where_clauses = ["status = 'ACTIVE'"]
    params = {}
    if req.sources:
        where_clauses.append("source = ANY(:sources)")
        params["sources"] = req.sources
    if req.states:
        where_clauses.append("state = ANY(:states)")
        params["states"] = req.states
    if req.authorities:
        where_clauses.append("organization = ANY(:authorities)")
        params["authorities"] = req.authorities
    if req.departments:
        where_clauses.append("department = ANY(:departments)")
        params["departments"] = req.departments
    if req.query:
        where_clauses.append("(title ILIKE :q OR description ILIKE :q)")
        params["q"] = f"%{req.query}%"
    where_sql = " AND ".join(where_clauses)
    result = await db.execute(text(f"SELECT COUNT(*) FROM tenders WHERE {where_sql}"), params)
    return {"count": result.scalar()}


@router.post("/export-xlsx")
async def export_filtered_xlsx(req: DownloadRequest, db: AsyncSession = Depends(get_db)):
    """Export filtered tenders as template-formatted XLSX matching the user's template."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    # Build query
    where_clauses = ["status = 'ACTIVE'"]
    params = {}

    if req.sources:
        where_clauses.append("source = ANY(:sources)")
        params["sources"] = req.sources
    if req.states:
        where_clauses.append("state = ANY(:states)")
        params["states"] = req.states
    if req.authorities:
        where_clauses.append("organization = ANY(:authorities)")
        params["authorities"] = req.authorities
    if req.departments:
        where_clauses.append("department = ANY(:departments)")
        params["departments"] = req.departments
    if req.query:
        where_clauses.append("(title ILIKE :q OR description ILIKE :q)")
        params["q"] = f"%{req.query}%"
    if req.min_value is not None:
        where_clauses.append("tender_value_estimated >= :min_val")
        params["min_val"] = float(req.min_value)
    if req.max_value is not None:
        where_clauses.append("tender_value_estimated <= :max_val")
        params["max_val"] = float(req.max_value)

    where_sql = " AND ".join(where_clauses)
    sql = text(
        f"SELECT tender_id, title, description, state, organization, department, "
        f"tender_value_estimated, emd_amount, publication_date, bid_open_date, "
        f"bid_close_date, source, source_url "
        f"FROM tenders WHERE {where_sql} "
        f"ORDER BY publication_date DESC NULLS LAST "
        f"LIMIT 2000"
    )
    result = await db.execute(sql, params)
    tenders = result.fetchall()

    # ── Build workbook matching template format ──
    wb = openpyxl.Workbook()
    ws = wb.active
    today_str = datetime.now().strftime("%d%m%Y")
    ws.title = f"Tender List_{today_str}"

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")
    currency_fmt = '#,##0'
    cr_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    # Template columns: S.NO, Location, Tender Authority, Summary,
    #   Tender Amount in Cr, Tender Amount, EMD, Opening Date, Closing Date,
    #   Status, Tender Id, Tender No
    headers = [
        "S.NO", "Location", "Tender Authority", "Summary",
        "Tender Amount in Cr", "Tender Amount", "EMD",
        "Opening Date", "Closing Date", "Status",
        "Tender Id", "Tender No",
    ]
    col_widths = [6, 22, 20, 58, 20, 16, 14, 14, 14, 12, 16, 16]

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write data rows
    today = date.today()
    for row_idx, t in enumerate(tenders, 2):
        (tender_id, title, description, state, organization, department,
         value_est, emd, pub_date, open_date, close_date, source, source_url) = t

        # Summary = title (or description if title is short)
        summary = title or ""
        if description and len(summary) < 50:
            summary = description[:500] if description else summary

        # Location = "State" or "Department, State"
        location_parts = []
        if state:
            location_parts.append(state)
        location = ", ".join(location_parts) if location_parts else (source or "").upper()

        # Tender amount: use estimated value, or EMD*50 fallback
        tender_amount = None
        if value_est:
            tender_amount = float(value_est)
        elif emd:
            tender_amount = float(emd) * 50  # same as template convention

        emd_val = float(emd) if emd else None

        # Status based on close date vs today
        status_str = "Open"
        if close_date:
            cd = close_date.date() if hasattr(close_date, 'date') else close_date
            if cd < today:
                status_str = "Closed"
            elif cd == today:
                status_str = "Closing Today"

        # S.NO
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = center_align

        # Location
        ws.cell(row=row_idx, column=2, value=location).alignment = data_align

        # Tender Authority
        ws.cell(row=row_idx, column=3, value=organization or department or "").alignment = data_align

        # Summary
        ws.cell(row=row_idx, column=4, value=summary).alignment = data_align

        # Tender Amount in Cr (formula)
        if tender_amount is not None:
            cell_cr = ws.cell(row=row_idx, column=5)
            cell_cr.value = tender_amount / 10000000  # convert to Cr
            cell_cr.number_format = cr_fmt
            cell_cr.alignment = center_align
        else:
            ws.cell(row=row_idx, column=5, value=None)

        # Tender Amount (raw)
        cell_amt = ws.cell(row=row_idx, column=6, value=tender_amount)
        if tender_amount:
            cell_amt.number_format = currency_fmt
        cell_amt.alignment = center_align

        # EMD
        cell_emd = ws.cell(row=row_idx, column=7, value=emd_val)
        if emd_val:
            cell_emd.number_format = currency_fmt
        cell_emd.alignment = center_align

        # Opening Date
        if open_date or pub_date:
            d = open_date or pub_date
            ws.cell(row=row_idx, column=8, value=d.strftime("%d %b %Y") if d else "").alignment = center_align
        else:
            ws.cell(row=row_idx, column=8, value="").alignment = center_align

        # Closing Date
        if close_date:
            ws.cell(row=row_idx, column=9, value=close_date.strftime("%d %b %Y")).alignment = center_align
        else:
            ws.cell(row=row_idx, column=9, value="").alignment = center_align

        # Status
        status_cell = ws.cell(row=row_idx, column=10, value=status_str)
        status_cell.alignment = center_align
        if status_str == "Open":
            status_cell.font = Font(color="228B22", bold=True)
        elif status_str == "Closed":
            status_cell.font = Font(color="CC0000")
        elif status_str == "Closing Today":
            status_cell.font = Font(color="FF8C00", bold=True)

        # Tender Id (internal / source)
        ws.cell(row=row_idx, column=11, value=tender_id or "").alignment = center_align

        # Tender No
        ws.cell(row=row_idx, column=12, value=tender_id or "").alignment = center_align

        # Borders + alternate row shading
        for col in range(1, 13):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Summary sheet (matching template Sheet1) ──
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=2, column=2, value="Tender List Export").font = Font(size=14, bold=True)
    ws2.cell(row=4, column=2, value="Generated:").font = Font(bold=True)
    ws2.cell(row=4, column=3, value=datetime.now().strftime("%d %b %Y, %I:%M %p"))
    ws2.cell(row=5, column=2, value="Total Tenders:").font = Font(bold=True)
    ws2.cell(row=5, column=3, value=len(tenders))

    # Filter summary
    row = 7
    ws2.cell(row=row, column=2, value="Applied Filters:").font = Font(size=11, bold=True, underline="single")
    row += 1
    if req.sources:
        ws2.cell(row=row, column=2, value="Sources:")
        ws2.cell(row=row, column=3, value=", ".join(s.upper() for s in req.sources))
        row += 1
    if req.states:
        ws2.cell(row=row, column=2, value="States:")
        ws2.cell(row=row, column=3, value=", ".join(req.states))
        row += 1
    if req.authorities:
        ws2.cell(row=row, column=2, value="Authorities:")
        ws2.cell(row=row, column=3, value=", ".join(req.authorities))
        row += 1
    if req.departments:
        ws2.cell(row=row, column=2, value="Departments:")
        ws2.cell(row=row, column=3, value=", ".join(req.departments))
        row += 1

    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 60

    # Total amount summary
    row += 1
    total_value = sum(
        (float(t[6]) if t[6] else (float(t[7]) * 50 if t[7] else 0))
        for t in tenders
    )
    ws2.cell(row=row, column=2, value="Total Tender Value:").font = Font(bold=True)
    cr_cell = ws2.cell(row=row, column=3, value=f"₹ {total_value / 10000000:,.2f} Cr")

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Tender_List_{today_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
